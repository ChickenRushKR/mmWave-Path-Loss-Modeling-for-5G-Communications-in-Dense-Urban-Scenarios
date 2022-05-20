import tensorflow as tf
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
# import glob
#import zipfile
from tensorflow import keras
# from tensorflow.keras import backend as K
# from tensorflow.keras import datasets, layers, models
from tensorflow.keras.layers import Input
from tensorflow.keras import Model
from tensorflow.keras.models import load_model
import argparse
# from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
import random
import tensorflow as tf
import os
# gpus = tf.config.experimental.list_physical_devices('GPU')
# if gpus:
#     try:
#         # Currently, memory growth needs to be the same across GPUs
#         for gpu in gpus:
#             tf.config.experimental.set_memory_growth(gpu, True)
#         logical_gpus = tf.config.experimental.list_logical_devices('GPU')
#         print(len(gpus), "Physical GPUs,", len(logical_gpus), "Logical GPUs")
#     except RuntimeError as e:
#         # Memory growth must be set before GPUs have been initialized
#         print(e)

def main(args):
    config = pd.read_csv(args.set)
    numeric_data_rss = pd.read_csv(args.csv)
    result = pd.read_excel('result.xlsx', engine='openpyxl')
    write_wb = load_workbook("result.xlsx")
    write_ws = write_wb.active
    write_ws['A1'] = 'idx'
    write_ws['B1'] = 'case'
    write_ws['C1'] = 'result'
    write_ws['D1'] = 'comment'
    write_ws['E1'] = 'time'
    write_ws['F1'] = 'cmd'
    write_ws['G1'] = 'lams'
    write_ws['H1'] = 'memo'
    line = len(result) + 2

    for n in range(args.n):
        for m in range(len(config)):
            train_test = config['train_test'][m]
            numeric_data_train = numeric_data_rss[numeric_data_rss[train_test] == 'train']
            numeric_data_test = numeric_data_rss[numeric_data_rss[train_test] == 'test']
            train_length = len(numeric_data_rss[numeric_data_rss[train_test] == 'train'])
            test_length = len(numeric_data_rss[numeric_data_rss[train_test] == 'test'])
            x_train = np.zeros((train_length, 2))
            x_test = np.zeros((test_length, 2))
            x_data = np.zeros((test_length + train_length, 2))

            train_lams_data = []
            test_lams_data = []
            train_lams_2ddata = []
            test_lams_2ddata = []
            lams_data = []
            lams_path = 'lams/' + config['lams'][m]
            lams_shape = config['lams'][m].split('_')
            lams_shape = [int(lams_shape[1]),int(lams_shape[2]),int(lams_shape[3])]
            # lams_2ddata = []
            # lams_2dpath = 'lams2d/' + config['lams2d'][m]
            # lams_2dshape = config['lams2d'][m].split('_')
            # lams_2dshape = [int(lams_2dshape[1]), int(lams_2dshape[2])]
            x = 0
            y = 0
            for i in range(len(numeric_data_rss)):
                lams_data.append(np.load(lams_path + '/' + str(i + 1) + '.npy'))
                x_data[x][0] = (180 - numeric_data_rss['chai1'][i]) / 180 * 20 # 25 -> 12
                x_data[x][1] = (262 - numeric_data_rss['dist'][i]) / 262

                if numeric_data_rss[train_test][i] == 'train':
                    train_lams_data.append(np.load(lams_path + '/' + str(i + 1) + '.npy'))
                    x_train[x][0] = (180 - numeric_data_rss['chai1'][i]) / 180 * 20
                    x_train[x][1] = (262 - numeric_data_rss['dist'][i]) / 262
                    x += 1
                else:
                    test_lams_data.append(np.load(lams_path + '/' + str(i + 1) + '.npy'))
                    x_test[y][0] = (180 - numeric_data_rss['chai1'][i]) / 180 * 20
                    x_test[y][1] = (262 - numeric_data_rss['dist'][i]) / 262
                    y += 1

            numeric_train = np.array(x_train).astype(np.float32)
            numeric_test = np.array(x_test).astype(np.float32)
            numeric_data = np.array(x_data).astype(np.float32)
            img_train = np.array(train_lams_data).astype(np.float32)
            img_test = np.array(test_lams_data).astype(np.float32)
            img_data = np.array(lams_data).astype(np.float32)

            labelA = numeric_data_rss[numeric_data_rss[train_test] == 'train']['RSRP'].to_numpy().astype(np.float32)
            labelB = numeric_data_rss[numeric_data_rss[train_test] == 'test']['RSRP'].to_numpy().astype(np.float32)
            labelC = numeric_data_rss['RSRP'].to_numpy().astype(np.float32)
            label_train = labelA
            label_train += 120
            label_test = labelB
            label_test += 120
            label_data = labelC
            label_data += 120

            img_train = np.expand_dims(img_train, axis=4)
            img_test = np.expand_dims(img_test, axis=4)
            img_data = np.expand_dims(img_data, axis=4)

            input_3dimg = Input(shape=(lams_shape[0], lams_shape[1], lams_shape[2], 1), dtype='float32')
            input_var1 = Input(shape=(1,), dtype='float32')
            input_var2 = Input(shape=(1,), dtype='float32')

            x = keras.layers.Conv3D(80, (4, 4, 4), strides=2, activation='relu', input_shape=(lams_shape[0], lams_shape[1], lams_shape[2], 1),
                                    padding='valid')(input_3dimg)
            x = keras.layers.Conv3D(20, (2, 2, 2), strides=2, activation='relu', padding='same')(x)
            x = keras.layers.Conv3D(5, (2, 2, 2), strides=1, activation='relu', padding='valid')(x)

            x = keras.layers.Flatten()(x)
            x = Model(inputs=input_3dimg, outputs=x)

            azimuth = keras.layers.Dense(1)(input_var1)
            azimuth = Model(inputs=input_var1, outputs=azimuth)
            dist = keras.layers.Dense(1)(input_var2)
            dist = Model(inputs=input_var2, outputs=dist)

            combined_3d = keras.layers.concatenate([x.output, azimuth.output, dist.output])
            combined_3d = keras.layers.Dense(10, activation='relu')(combined_3d)
            combined_3d = keras.layers.Dense(5, activation='relu')(combined_3d)
            combined_3d = keras.layers.Dense(1, activation='relu')(combined_3d)
            model = Model(inputs=[x.input, input_var1, input_var2], outputs=combined_3d)

            RMSE = keras.metrics.RootMeanSquaredError()
            MAE = keras.metrics.MeanAbsoluteError()

            opt = keras.optimizers.Adam(learning_rate=args.lr)
            model.compile(optimizer=opt, loss='mean_squared_error', metrics=[RMSE, MAE])

            tasks = [[2, 4, 5, 6], [1, 2, 3, 4], [1, 2, 3, 4]]
            cases = ['NR Serving', 'dist_level', 'chai_level']
            flag = True
            random_prev = [['NR Serving', 2], ['NR Serving', 4], ['NR Serving', 5], ['NR Serving', 6], ['dist_level', 1], ['dist_level', 2],
                           ['dist_level', 3], ['dist_level', 4], ['chai_level', 1], ['chai_level', 2], ['chai_level', 3], ['chai_level', 4]]

            for epoch in range(args.outer):
                if flag is False:
                    break
                print("###################### epoch: ", epoch + 1, " / ", args.outer,"######################")
                for _ in range(4):
                    print("batch: ", _ + 1, "/", 12, " case: ", random_prev[(epoch * 4 + _) % 12][0], " task: ", random_prev[(epoch * 4 + _) % 12][1])
                    x = []
                    x2 = []
                    x3 = []
                    y = []
                    for idx in numeric_data_train[numeric_data_train[random_prev[(epoch * 4 + _) % 12][0]] == random_prev[(epoch * 4 + _) % 12][1]]['idx']:
                        x.append(img_data[idx - 1])
                        x2.append(numeric_data[idx - 1][0])
                        x3.append(numeric_data[idx - 1][1])
                        y.append(label_data[idx - 1])
                    x_np = np.array(x)
                    x2_np = np.array(x2)
                    x3_np = np.array(x3)
                    y_np = np.array(y)
                    if epoch%3 == 0:
                        azimuth.trainable = False
                        dist.trainable = False
                        model.compile(optimizer=opt, loss='mean_squared_error', metrics=[RMSE, MAE])
                    elif epoch%3 == 1:
                        dist.trainable = True
                        model.compile(optimizer=opt, loss='mean_squared_error', metrics=[RMSE, MAE])
                    else:
                        dist.trainable = False
                        azimuth.trainable = True
                        model.compile(optimizer=opt, loss='mean_squared_error', metrics=[RMSE, MAE])
                    history = model.fit([x_np, x2_np, x3_np], y_np, epochs=args.inner, verbose=0)
                    print("RMSE: ", round(history.history['root_mean_squared_error'][args.inner - 1], 2))
                    if round(history.history['root_mean_squared_error'][args.inner - 1], 2) > args.threshold and args.nostop == 1:
                        flag = False
                        break
                ev = model.evaluate([img_test, numeric_test[:,0], numeric_test[:,1]], label_test, verbose=0)
            if flag is False:
                if m > 0:
                    m -= 1
                else:
                    m = 0
                continue

            predictions = model.predict([img_test,  numeric_test[:,0], numeric_test[:,1]])
            # predictions_train = model.predict([img_train, numeric_train])
            evaluate = model.evaluate([img_test,  numeric_test[:,0], numeric_test[:,1]], label_test, verbose=2)
            # evaluate2_ = model.evaluate([img_train, numeric_train], label_train, verbose=2)
            rmse = str(round(evaluate[1], 4))
            # rmse2_ = str(round(evaluate2_[1], 4))
            comment = '3D_' + '(' + train_test + ')_' + str(line)

            if float(rmse) < 9.0:
                for i in range(0, len(label_test) - 1):
                    plt.plot(i, predictions[i][0], c='red', marker='^', markersize=4)
                    plt.plot(i, label_test[i], c='blue', marker='^', markersize=4)
                plt.xlabel("index")
                plt.ylabel("RSRP (+120dBm)")
                plt.title('RMSE: ' + rmse)
                plt.legend(['Predict', 'Measured'])
                # plt.show()

                plt.savefig('result/' + comment + '.png')
                plt.cla()

                # for i in range(0, len(label_train) - 1):
                #     plt.plot(i, predictions_train[i][0], c='red', marker='^', markersize=4)
                #     plt.plot(i, label_train[i], c='blue', marker='^', markersize=4)
                # plt.xlabel("index")
                # plt.ylabel("RSRP (+120dBm)")
                # plt.title('RMSE: ' + rmse2_)
                # plt.legend(['Predict', 'Measured'])
                # plt.show()
                #
                # plt.savefig('result/' + comment + '(train).png')
                # plt.cla()
                #
                # A = [[]]
                # for i in range(len(label_test)):
                #     A.append([(262 - (numeric_test[i][2]*262)), label_test[i], predictions[i][0]])
                #
                # # print(A)
                # A = A[1:]
                # A.sort(key=lambda x: x[0])
                #
                # # print(A)
                # for i in range(len(label_test)):
                #     plt.plot(A[i][0], A[i][2], c='red', marker='^', markersize=4)
                #     plt.plot(A[i][0], A[i][1], c='blue', marker='^', markersize=4)  # label
                # plt.xlabel("Distance")
                # plt.ylabel("RSRP (+120dBm)")
                # plt.title('RMSE: ' + rmse)
                # plt.legend(['Predict', 'Measured'])
                # plt.xlim([0, 280])
                # plt.ylim([0, 50])
                # plt.savefig('result/' + comment + '(distance).png')
                # plt.cla()

                model.save("saved/" + str(line) + "_" + rmse + ".h5")

            now = datetime.datetime.now()
            write_ws.cell(line, 1, line)
            write_ws.cell(line, 2, train_test)
            write_ws.cell(line, 3, evaluate[1])
            write_ws.cell(line, 4, comment)
            write_ws.cell(line, 5, now.strftime('%Y-%m-%d %H:%M:%S'))
            write_ws.cell(line, 6, str(args))
            write_ws.cell(line, 7, config['lams'][m])
            write_ws.cell(line, 8, args.memo)

            line += 1
            write_wb.save("result.xlsx")
            keras.backend.clear_session()


if __name__ == '__main__':
    # physical_devices = tf.config.experimental.list_physical_devices('GPU')
    # tf.config.experimental.set_memory_growth(physical_devices[0], True)
    parser = argparse.ArgumentParser()
    parser.add_argument('-n', required=False, type=int, default=10, help='실행 횟수')
    parser.add_argument('-inner', required=False, type=int, default=40, help='inner 횟수')
    parser.add_argument('-outer', required=False, type=int, default=1, help='iterator 횟수')
    parser.add_argument('-batch_size', required=False, type=int, default=10, help='batch_size')
    parser.add_argument('-ft', required=False, type=int, default=0, help='fine-tuning 횟수')
    parser.add_argument('-threshold', required=False, type=int, default=20, help='RMSE가 threshold넘으면 멈추기')
    parser.add_argument('-csv', required=False, type=str, default='RSS_meta_batch.csv', help='RSS 파일')
    parser.add_argument('-nostop', required=False, type=int, default=1, help='stop: 0, no stop: 1')
    parser.add_argument('-set', '-setting', required=False, type=str, default='setting.csv', help='설정 파일')
    parser.add_argument('-memo', required=False, type=str, default=' ', help='메모')
    parser.add_argument('-lr', required=False, type=float, default=0.001, help='learning rate')
    args = parser.parse_args()
    main(args)
